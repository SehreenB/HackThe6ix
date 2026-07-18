import argparse
from pathlib import Path

import requests
import openpyxl

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

SOURCES = {
    "ed_visits_2024_2025_final": (
        "https://www.cihi.ca/sites/default/files/document/"
        "emergency-department-visits-apr-2024-mar-2025-data-tables-en.xlsx"
    ),
    "ed_visits_2025_2026_provisional": (
        "https://www.cihi.ca/sites/default/files/document/"
        "emergency-department-visits-apr-sep-2025-provisional-data-tables-en.xlsx"
    ),
}


def download_all(raw_dir: Path):
    for key, url in SOURCES.items():
        out_path = raw_dir / f"{key}.xlsx"
        if out_path.exists():
            print(f"{key}: already downloaded, skipping")
            continue
        print(f"Downloading {key} ...")
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        out_path.write_bytes(resp.content)
        print(f"  saved {len(resp.content):,} bytes to {out_path}")


def inspect_all(raw_dir: Path):
    for key in SOURCES:
        path = raw_dir / f"{key}.xlsx"
        if not path.exists():
            print(f"{key}: not downloaded yet")
            continue
        print(f"\n=== {key} ===")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Sheet: {sheet_name}  ({ws.max_row} rows x {ws.max_column} cols)")
            # print first couple of rows so you can see actual column headers
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
                print(f"    row {i}: {row}")


import json
import re

def clean_text(text):
    if text is None:
        return None
    return re.sub(r'[*†‡§]+', '', str(text)).strip()

def extract_json(raw_dir: Path):
    path = raw_dir / "ed_visits_2024_2025_final.xlsx"
    if not path.exists():
        print(f"Data file not found at {path}")
        return

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Table 1" not in wb.sheetnames:
        print("Table 1 not found in workbook")
        return

    ws = wb["Table 1"]

    # Canonical PRUIDs to ensure we don't silently drop any province
    ALL_PROVS = {
        "N.L.": "NL", "Newfoundland and Labrador": "NL", "Newfoundland": "NL",
        "P.E.I.": "PE", "Prince Edward Island": "PE",
        "N.S.": "NS", "Nova Scotia": "NS",
        "N.B.": "NB", "New Brunswick": "NB",
        "Que.": "QC", "Quebec": "QC",
        "Ont.": "ON", "Ontario": "ON",
        "Man.": "MB", "Manitoba": "MB",
        "Sask.": "SK", "Saskatchewan": "SK",
        "Alta.": "AB", "Alberta": "AB",
        "B.C.": "BC", "British Columbia": "BC",
        "Y.T.": "YT", "Yukon": "YT",
        "N.W.T.": "NT", "Northwest Territories": "NT",
        "Nun.": "NU", "Nunavut": "NU"
    }

    results = {
        code: {
            "ed_visits_total": "data not available",
            "median_los_admitted_hours": "data not available",
            "p90_los_admitted_hours": "data not available"
        }
        for code in set(ALL_PROVS.values())
    }

    def parse_val(val):
        if val is None or str(val).strip() in ('', '.', '..', 'x', 'X', '-'):
            return "data not available"
        return val

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx < 2:  # Skip title and headers
            continue
            
        raw_prov = row[0]
        if not raw_prov:
            continue
            
        prov = clean_text(raw_prov)
        if prov.startswith("Notes") or prov.startswith("Source") or "Table" in prov:
            break

        pruid = ALL_PROVS.get(prov)
        if not pruid:
            continue

        results[pruid] = {
            "ed_visits_total": parse_val(row[1]),
            "median_los_admitted_hours": parse_val(row[7]),
            "p90_los_admitted_hours": parse_val(row[10])
        }

    out_path = DATA_DIR / "cihi_provincial_metrics.json"
    out_path.write_text(json.dumps(results, indent=2))
    print(f"\nExtracted data for {len(results)} provinces/territories to {out_path.name}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--inspect", action="store_true",
                         help="Download (if needed) and print sheet/column structure, don't extract JSON yet")
    parser.add_argument("--raw-dir", default=str(DATA_DIR / "_raw_cihi"))
    args = parser.parse_args()

    raw_dir = Path(args.raw_dir)
    raw_dir.mkdir(parents=True, exist_ok=True)

    download_all(raw_dir)

    if args.inspect:
        inspect_all(raw_dir)
        print("\nInspect the sheet names and header rows above, then update this "
              "script's extraction logic to pull the specific province-level "
              "columns you need (visit volume, median LOS, 90th percentile LOS) "
              "before wiring the result into investment.py.")
        return

    extract_json(raw_dir)
if __name__ == "__main__":
    main()
